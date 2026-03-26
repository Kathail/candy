import csv
import io
from collections import defaultdict
from datetime import datetime, timedelta, timezone

from flask import Blueprint, Response, render_template, request
from flask_login import login_required

from app import db
from app.helpers import sanitize_csv_value as _s
from app.models import Customer, Payment

bp = Blueprint("reports", __name__)


@bp.route("/reports")
@login_required
def reports():
    """Reports and exports page for tax purposes"""
    today = datetime.now(timezone.utc).date()

    # Calculate date ranges
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)

    month_start = today.replace(day=1)
    if today.month == 12:
        month_end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        month_end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)

    quarter = (today.month - 1) // 3
    quarter_start = today.replace(month=quarter * 3 + 1, day=1)
    if quarter == 3:
        quarter_end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        quarter_end = today.replace(month=(quarter + 1) * 3 + 1, day=1) - timedelta(days=1)

    year_start = today.replace(month=1, day=1)
    year_end = today.replace(month=12, day=31)

    def get_period_stats(start, end):
        payments = Payment.query.filter(
            Payment.payment_date >= start,
            Payment.payment_date <= end
        ).all()
        total = sum(p.amount for p in payments)
        count = len(payments)
        return {"total": total, "count": count}

    week_stats = get_period_stats(week_start, week_end)
    month_stats = get_period_stats(month_start, month_end)
    quarter_stats = get_period_stats(quarter_start, quarter_end)
    year_stats = get_period_stats(year_start, year_end)

    # Get tax-exempt stats
    tax_exempt_customers = Customer.query.filter_by(tax_exempt=True, status='active').all()
    tax_exempt_ids = [c.id for c in tax_exempt_customers]

    def get_tax_exempt_stats(start, end):
        if not tax_exempt_ids:
            return {"total": 0, "count": 0, "payments": []}
        payments = Payment.query.filter(
            Payment.customer_id.in_(tax_exempt_ids),
            Payment.payment_date >= start,
            Payment.payment_date <= end
        ).order_by(Payment.payment_date.desc()).all()
        total = sum(p.amount for p in payments)
        count = len(payments)
        return {"total": total, "count": count, "payments": payments}

    tax_exempt_week = get_tax_exempt_stats(week_start, week_end)
    tax_exempt_month = get_tax_exempt_stats(month_start, month_end)
    tax_exempt_quarter = get_tax_exempt_stats(quarter_start, quarter_end)
    tax_exempt_year = get_tax_exempt_stats(year_start, year_end)

    return render_template(
        "reports.html",
        today=today,
        week_start=week_start,
        week_end=week_end,
        month_start=month_start,
        month_end=month_end,
        quarter_start=quarter_start,
        quarter_end=quarter_end,
        year_start=year_start,
        year_end=year_end,
        week_stats=week_stats,
        month_stats=month_stats,
        quarter_stats=quarter_stats,
        year_stats=year_stats,
        tax_exempt_customers=tax_exempt_customers,
        tax_exempt_week=tax_exempt_week,
        tax_exempt_month=tax_exempt_month,
        tax_exempt_quarter=tax_exempt_quarter,
        tax_exempt_year=tax_exempt_year,
    )


@bp.route("/reports/export")
@login_required
def export_report():
    """Export financial report as CSV or PDF"""
    report_type = request.args.get("type", "payments")
    format_type = request.args.get("format", "csv")
    start_date = request.args.get("start")
    end_date = request.args.get("end")
    period = request.args.get("period")

    today = datetime.now(timezone.utc).date()

    # Calculate date range based on period
    if period == "week":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
    elif period == "month":
        start = today.replace(day=1)
        if today.month == 12:
            end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    elif period == "quarter":
        quarter = (today.month - 1) // 3
        start = today.replace(month=quarter * 3 + 1, day=1)
        if quarter == 3:
            end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end = today.replace(month=(quarter + 1) * 3 + 1, day=1) - timedelta(days=1)
    elif period == "year":
        start = today.replace(month=1, day=1)
        end = today.replace(month=12, day=31)
    elif start_date and end_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            start = None
            end = None
    else:
        start = None
        end = None

    # Get data based on report type
    if report_type == "payments":
        query = Payment.query.options(db.joinedload(Payment.customer))
        if start and end:
            query = query.filter(Payment.payment_date >= start, Payment.payment_date <= end)
        payments = query.order_by(Payment.payment_date.desc()).all()

        if format_type == "pdf":
            return _generate_payments_pdf(payments, start, end)
        else:
            return _generate_payments_csv(payments, start, end)

    elif report_type == "summary":
        query = Payment.query.options(db.joinedload(Payment.customer))
        if start and end:
            query = query.filter(Payment.payment_date >= start, Payment.payment_date <= end)
        payments = query.order_by(Payment.payment_date.desc()).all()

        if format_type == "pdf":
            return _generate_summary_pdf(payments, start, end)
        else:
            return _generate_summary_csv(payments, start, end)

    elif report_type == "balances":
        customers = Customer.query.filter(
            Customer.balance > 0,
            Customer.status == 'active'
        ).order_by(Customer.balance.desc()).all()

        if format_type == "pdf":
            return _generate_balances_pdf(customers, start, end)
        else:
            return _generate_balances_csv(customers)

    elif report_type == "tax_exempt":
        tax_exempt_customers = Customer.query.filter_by(tax_exempt=True, status='active').all()
        tax_exempt_ids = [c.id for c in tax_exempt_customers]

        query = Payment.query.options(db.joinedload(Payment.customer)).filter(
            Payment.customer_id.in_(tax_exempt_ids) if tax_exempt_ids else False
        )
        if start and end:
            query = query.filter(Payment.payment_date >= start, Payment.payment_date <= end)
        payments = query.order_by(Payment.payment_date.desc()).all()

        if format_type == "pdf":
            return _generate_tax_exempt_pdf(payments, tax_exempt_customers, start, end)
        elif format_type == "excel":
            return _generate_tax_exempt_excel(payments, tax_exempt_customers, start, end)
        else:
            return _generate_tax_exempt_csv(payments, tax_exempt_customers, start, end)

    return "Invalid report type", 400


# --- Report generation helpers ---

def _generate_payments_csv(payments, start, end):
    output = io.StringIO()
    writer = csv.writer(output)

    period_str = f"{start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}" if start and end else "All Time"
    writer.writerow([f"Payment Report - {period_str}"])
    writer.writerow([])
    writer.writerow(["Date", "Receipt #", "Customer", "City", "Amount", "Previous Balance", "Notes"])

    total = 0
    for p in payments:
        writer.writerow([
            p.payment_date.strftime("%Y-%m-%d"),
            p.receipt_number or "",
            _s(p.customer.name if p.customer else ""),
            _s(p.customer.city if p.customer else ""),
            f"{p.amount:.2f}",
            f"{p.previous_balance:.2f}" if p.previous_balance else "",
            _s(p.notes or "")
        ])
        total += float(p.amount)

    writer.writerow([])
    writer.writerow(["", "", "", "TOTAL:", f"${total:.2f}", "", ""])

    output.seek(0)
    filename = f"payments_{start.strftime('%Y%m%d') if start else 'all'}_{end.strftime('%Y%m%d') if end else 'time'}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def _generate_payments_pdf(payments, start, end):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=10, alignment=1)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=12, spaceAfter=20, alignment=1)

    elements.append(Paragraph("Payment Report", title_style))
    period_str = f"{start.strftime('%B %d, %Y')} to {end.strftime('%B %d, %Y')}" if start and end else "All Time"
    elements.append(Paragraph(period_str, subtitle_style))
    elements.append(Spacer(1, 20))

    total = sum(p.amount for p in payments)
    elements.append(Paragraph(f"<b>Total Payments:</b> {len(payments)}", styles['Normal']))
    elements.append(Paragraph(f"<b>Total Collected:</b> ${total:.2f}", styles['Normal']))
    elements.append(Spacer(1, 20))

    if payments:
        table_data = [["Date", "Receipt #", "Customer", "City", "Amount"]]
        for p in payments:
            table_data.append([
                p.payment_date.strftime("%m/%d/%Y"),
                p.receipt_number or "-",
                (p.customer.name if p.customer else "-")[:25],
                (p.customer.city if p.customer else "-")[:15],
                f"${p.amount:.2f}"
            ])
        table_data.append(["", "", "", "TOTAL:", f"${total:.2f}"])

        table = Table(table_data, colWidths=[1*inch, 1*inch, 2.5*inch, 1.5*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.6)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"payments_{start.strftime('%Y%m%d') if start else 'all'}_{end.strftime('%Y%m%d') if end else 'time'}.pdf"
    return Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def _generate_summary_csv(payments, start, end):
    output = io.StringIO()
    writer = csv.writer(output)

    period_str = f"{start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}" if start and end else "All Time"
    writer.writerow([f"Financial Summary Report - {period_str}"])
    writer.writerow([])

    total = sum(p.amount for p in payments)
    writer.writerow(["SUMMARY"])
    writer.writerow(["Total Payments", len(payments)])
    writer.writerow(["Total Collected", f"${total:.2f}"])
    writer.writerow(["Average Payment", f"${(total/len(payments)):.2f}" if payments else "$0.00"])
    writer.writerow([])

    by_customer = defaultdict(lambda: {"count": 0, "total": 0})
    for p in payments:
        name = p.customer.name if p.customer else "Unknown"
        by_customer[name]["count"] += 1
        by_customer[name]["total"] += float(p.amount)

    writer.writerow(["BY CUSTOMER"])
    writer.writerow(["Customer", "Payment Count", "Total Amount"])
    for name, data in sorted(by_customer.items(), key=lambda x: x[1]["total"], reverse=True):
        writer.writerow([_s(name), data["count"], f"${data['total']:.2f}"])
    writer.writerow([])

    by_city = defaultdict(lambda: {"count": 0, "total": 0})
    for p in payments:
        city = p.customer.city if p.customer and p.customer.city else "Unknown"
        by_city[city]["count"] += 1
        by_city[city]["total"] += float(p.amount)

    writer.writerow(["BY CITY"])
    writer.writerow(["City", "Payment Count", "Total Amount"])
    for city, data in sorted(by_city.items(), key=lambda x: x[1]["total"], reverse=True):
        writer.writerow([_s(city), data["count"], f"${data['total']:.2f}"])
    writer.writerow([])

    by_month = defaultdict(lambda: {"count": 0, "total": 0})
    for p in payments:
        month_key = p.payment_date.strftime("%Y-%m")
        by_month[month_key]["count"] += 1
        by_month[month_key]["total"] += float(p.amount)

    writer.writerow(["BY MONTH"])
    writer.writerow(["Month", "Payment Count", "Total Amount"])
    for month, data in sorted(by_month.items()):
        writer.writerow([month, data["count"], f"${data['total']:.2f}"])

    output.seek(0)
    filename = f"summary_{start.strftime('%Y%m%d') if start else 'all'}_{end.strftime('%Y%m%d') if end else 'time'}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def _generate_summary_pdf(payments, start, end):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=10, alignment=1)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=12, spaceAfter=20, alignment=1)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=14, spaceBefore=20, spaceAfter=10)

    elements.append(Paragraph("Financial Summary Report", title_style))
    period_str = f"{start.strftime('%B %d, %Y')} to {end.strftime('%B %d, %Y')}" if start and end else "All Time"
    elements.append(Paragraph(period_str, subtitle_style))
    elements.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%B %d, %Y at %I:%M %p')}", subtitle_style))
    elements.append(Spacer(1, 20))

    total = sum(p.amount for p in payments)
    avg = total / len(payments) if payments else 0

    summary_data = [
        ["Total Payments", str(len(payments))],
        ["Total Collected", f"${total:.2f}"],
        ["Average Payment", f"${avg:.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # By Customer
    elements.append(Paragraph("Collections by Customer", section_style))
    by_customer = defaultdict(lambda: {"count": 0, "total": 0})
    for p in payments:
        name = p.customer.name if p.customer else "Unknown"
        by_customer[name]["count"] += 1
        by_customer[name]["total"] += float(p.amount)

    customer_data = [["Customer", "Payments", "Total"]]
    for name, data in sorted(by_customer.items(), key=lambda x: x[1]["total"], reverse=True)[:15]:
        customer_data.append([name[:30], str(data["count"]), f"${data['total']:.2f}"])

    if customer_data:
        customer_table = Table(customer_data, colWidths=[3.5*inch, 1*inch, 1.5*inch])
        customer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.6)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(customer_table)

    # By Month
    elements.append(Paragraph("Collections by Month", section_style))
    by_month = defaultdict(lambda: {"count": 0, "total": 0})
    for p in payments:
        month_key = p.payment_date.strftime("%B %Y")
        by_month[month_key]["count"] += 1
        by_month[month_key]["total"] += float(p.amount)

    month_data = [["Month", "Payments", "Total"]]
    for month, data in sorted(by_month.items(), key=lambda x: x[0]):
        month_data.append([month, str(data["count"]), f"${data['total']:.2f}"])

    if len(month_data) > 1:
        month_table = Table(month_data, colWidths=[3.5*inch, 1*inch, 1.5*inch])
        month_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.6)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(month_table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"summary_{start.strftime('%Y%m%d') if start else 'all'}_{end.strftime('%Y%m%d') if end else 'time'}.pdf"
    return Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def _generate_balances_csv(customers):
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([f"Outstanding Balances Report - {datetime.now(timezone.utc).strftime('%Y-%m-%d')}"])
    writer.writerow([])
    writer.writerow(["Customer", "City", "Phone", "Balance", "Last Visit"])

    total = 0
    for c in customers:
        writer.writerow([
            _s(c.name),
            _s(c.city or ""),
            _s(c.phone or ""),
            f"${c.balance:.2f}",
            c.last_visit.strftime("%Y-%m-%d") if c.last_visit else "Never"
        ])
        total += float(c.balance)

    writer.writerow([])
    writer.writerow(["", "", "TOTAL:", f"${total:.2f}", ""])

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=\"balances_{datetime.now(timezone.utc).strftime('%Y%m%d')}.csv\""}
    )


def _generate_balances_pdf(customers, start, end):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=10, alignment=1)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=12, spaceAfter=20, alignment=1)

    elements.append(Paragraph("Outstanding Balances Report", title_style))
    elements.append(Paragraph(f"As of {datetime.now(timezone.utc).strftime('%B %d, %Y')}", subtitle_style))
    elements.append(Spacer(1, 20))

    total = sum(c.balance for c in customers)
    elements.append(Paragraph(f"<b>Total Customers with Balance:</b> {len(customers)}", styles['Normal']))
    elements.append(Paragraph(f"<b>Total Outstanding:</b> ${total:.2f}", styles['Normal']))
    elements.append(Spacer(1, 20))

    if customers:
        table_data = [["Customer", "City", "Phone", "Balance", "Last Visit"]]
        for c in customers:
            table_data.append([
                c.name[:25],
                (c.city or "-")[:15],
                c.phone or "-",
                f"${c.balance:.2f}",
                c.last_visit.strftime("%m/%d/%Y") if c.last_visit else "Never"
            ])
        table_data.append(["", "", "TOTAL:", f"${total:.2f}", ""])

        table = Table(table_data, colWidths=[2*inch, 1.2*inch, 1.2*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.6, 0.2, 0.2)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=\"balances_{datetime.now(timezone.utc).strftime('%Y%m%d')}.pdf\""}
    )


def _generate_tax_exempt_csv(payments, customers, start, end):
    output = io.StringIO()
    writer = csv.writer(output)

    period_str = f"{start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}" if start and end else "All Time"
    writer.writerow([f"Tax Exempt Sales Report - {period_str}"])
    writer.writerow([])

    writer.writerow(["SUMMARY"])
    writer.writerow(["Tax Exempt Customers:", len(customers)])
    writer.writerow(["Total Transactions:", len(payments)])
    total = sum(p.amount for p in payments)
    writer.writerow(["Total Tax Exempt Sales:", f"${total:.2f}"])
    writer.writerow([])

    writer.writerow(["TAX EXEMPT CUSTOMERS"])
    writer.writerow(["Name", "City", "Phone", "Address"])
    for c in customers:
        writer.writerow([_s(c.name), _s(c.city or ""), _s(c.phone or ""), _s(c.address or "")])
    writer.writerow([])

    writer.writerow(["TRANSACTION DETAILS"])
    writer.writerow(["Date", "Receipt #", "Customer", "City", "Amount", "Notes"])
    for p in payments:
        writer.writerow([
            p.payment_date.strftime("%Y-%m-%d"),
            p.receipt_number or "",
            _s(p.customer.name if p.customer else ""),
            _s(p.customer.city if p.customer else ""),
            f"${p.amount:.2f}",
            _s(p.notes or "")
        ])
    writer.writerow([])
    writer.writerow(["", "", "", "TOTAL:", f"${total:.2f}", ""])

    output.seek(0)
    filename = f"tax_exempt_{start.strftime('%Y%m%d') if start else 'all'}_{end.strftime('%Y%m%d') if end else 'time'}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def _generate_tax_exempt_pdf(payments, customers, start, end):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=10, alignment=1)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=12, spaceAfter=20, alignment=1)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=14, spaceAfter=10, spaceBefore=20)

    elements.append(Paragraph("Tax Exempt Sales Report", title_style))
    period_str = f"{start.strftime('%B %d, %Y')} to {end.strftime('%B %d, %Y')}" if start and end else "All Time"
    elements.append(Paragraph(period_str, subtitle_style))
    elements.append(Spacer(1, 10))

    total = sum(p.amount for p in payments)
    elements.append(Paragraph(f"<b>Tax Exempt Customers:</b> {len(customers)}", styles['Normal']))
    elements.append(Paragraph(f"<b>Total Transactions:</b> {len(payments)}", styles['Normal']))
    elements.append(Paragraph(f"<b>Total Tax Exempt Sales:</b> ${total:.2f}", styles['Normal']))
    elements.append(Spacer(1, 20))

    if customers:
        elements.append(Paragraph("Tax Exempt Customers", section_style))
        cust_data = [["Name", "City", "Phone"]]
        for c in customers:
            cust_data.append([c.name[:30], (c.city or "-")[:20], c.phone or "-"])

        cust_table = Table(cust_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
        cust_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.8, 0.6, 0.2)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(cust_table)
        elements.append(Spacer(1, 20))

    if payments:
        elements.append(Paragraph("Transaction Details", section_style))
        table_data = [["Date", "Receipt #", "Customer", "Amount"]]
        for p in payments:
            table_data.append([
                p.payment_date.strftime("%m/%d/%Y"),
                p.receipt_number or "-",
                (p.customer.name if p.customer else "-")[:25],
                f"${p.amount:.2f}"
            ])
        table_data.append(["", "", "TOTAL:", f"${total:.2f}"])

        table = Table(table_data, colWidths=[1.2*inch, 1.2*inch, 2.5*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.8, 0.6, 0.2)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"tax_exempt_{start.strftime('%Y%m%d') if start else 'all'}_{end.strftime('%Y%m%d') if end else 'time'}.pdf"
    return Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def _generate_tax_exempt_excel(payments, customers, start, end):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_fill = PatternFill(start_color="D4A84B", end_color="D4A84B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    period_str = f"{start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}" if start and end else "All Time"
    ws_summary['A1'] = "Tax Exempt Sales Report"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A2'] = period_str
    ws_summary['A2'].font = Font(size=12, italic=True)

    ws_summary['A4'] = "Tax Exempt Customers:"
    ws_summary['B4'] = len(customers)
    ws_summary['A5'] = "Total Transactions:"
    ws_summary['B5'] = len(payments)
    total = sum(p.amount for p in payments)
    ws_summary['A6'] = "Total Tax Exempt Sales:"
    ws_summary['B6'] = f"${total:.2f}"

    for cell in ['A4', 'A5', 'A6']:
        ws_summary[cell].font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20

    # Customers sheet
    ws_customers = wb.create_sheet("Tax Exempt Customers")
    headers = ["Name", "City", "Phone", "Address"]
    for col, header in enumerate(headers, 1):
        cell = ws_customers.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row, c in enumerate(customers, 2):
        ws_customers.cell(row=row, column=1, value=c.name).border = thin_border
        ws_customers.cell(row=row, column=2, value=c.city or "").border = thin_border
        ws_customers.cell(row=row, column=3, value=c.phone or "").border = thin_border
        ws_customers.cell(row=row, column=4, value=c.address or "").border = thin_border

    ws_customers.column_dimensions['A'].width = 30
    ws_customers.column_dimensions['B'].width = 20
    ws_customers.column_dimensions['C'].width = 15
    ws_customers.column_dimensions['D'].width = 40

    # Transactions sheet
    ws_trans = wb.create_sheet("Transactions")
    headers = ["Date", "Receipt #", "Customer", "City", "Amount", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws_trans.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row, p in enumerate(payments, 2):
        ws_trans.cell(row=row, column=1, value=p.payment_date.strftime("%Y-%m-%d")).border = thin_border
        ws_trans.cell(row=row, column=2, value=p.receipt_number or "").border = thin_border
        ws_trans.cell(row=row, column=3, value=p.customer.name if p.customer else "").border = thin_border
        ws_trans.cell(row=row, column=4, value=p.customer.city if p.customer else "").border = thin_border
        ws_trans.cell(row=row, column=5, value=p.amount).border = thin_border
        ws_trans.cell(row=row, column=5).number_format = '$#,##0.00'
        ws_trans.cell(row=row, column=6, value=p.notes or "").border = thin_border

    total_row = len(payments) + 2
    ws_trans.cell(row=total_row, column=4, value="TOTAL:").font = Font(bold=True)
    ws_trans.cell(row=total_row, column=5, value=total).font = Font(bold=True)
    ws_trans.cell(row=total_row, column=5).number_format = '$#,##0.00'

    ws_trans.column_dimensions['A'].width = 12
    ws_trans.column_dimensions['B'].width = 12
    ws_trans.column_dimensions['C'].width = 25
    ws_trans.column_dimensions['D'].width = 15
    ws_trans.column_dimensions['E'].width = 12
    ws_trans.column_dimensions['F'].width = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"tax_exempt_{start.strftime('%Y%m%d') if start else 'all'}_{end.strftime('%Y%m%d') if end else 'time'}.xlsx"
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
